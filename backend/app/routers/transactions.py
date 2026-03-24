from fastapi import APIRouter, HTTPException, Query, Response
from bson import ObjectId
from pydantic import BaseModel
from typing import Optional
from ..database import get_db
from ..services.expense_classifier import EXPENSE_TAXONOMY, classify_expense, is_valid_kategorie

router = APIRouter(prefix="/api/transactions", tags=["transactions"])


def _fix_id(doc: dict) -> dict:
    doc["_id"] = str(doc["_id"])
    return doc


@router.get("/")
async def list_transactions(
    account_number: Optional[str] = None,
    transaction_type: Optional[str] = None,
    owner_unit: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    search: Optional[str] = None,
    haupttyp: Optional[str] = None,
    untertyp: Optional[str] = None,
    limit: int = Query(default=200, le=1000),
    skip: int = 0,
):
    db = get_db()
    query: dict = {}
    if account_number:
        query["account_number"] = account_number
    if transaction_type:
        query["transaction_type"] = transaction_type
    if owner_unit:
        query["owner_unit"] = owner_unit
    if haupttyp:
        query["haupttyp"] = haupttyp
    if untertyp:
        query["untertyp"] = untertyp
    if search:
        import re
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query["$or"] = [
            {"counterparty_name": {"$regex": pattern}},
            {"purpose": {"$regex": pattern}},
            {"booking_text": {"$regex": pattern}},
        ]
    if year:
        from datetime import date
        date_from = date(year, month or 1, 1).isoformat()
        if month:
            from calendar import monthrange
            last_day = monthrange(year, month)[1]
            date_to = date(year, month, last_day).isoformat()
        else:
            date_to = date(year, 12, 31).isoformat()
        query["booking_date"] = {"$gte": date_from, "$lte": date_to}

    total = await db.transactions.count_documents(query)
    docs = (
        await db.transactions.find(query)
        .sort("booking_date", -1)
        .skip(skip)
        .limit(limit)
        .to_list(None)
    )
    return {"total": total, "items": [_fix_id(d) for d in docs]}


# ---------------------------------------------------------------------------
# Expense taxonomy: used by frontend to build dropdowns
# Static path – MUST be defined before the /{tx_id} wildcard route
# ---------------------------------------------------------------------------

@router.get("/taxonomy")
async def get_taxonomy():
    """Return the full Haupttyp → [Untertypen] taxonomy for expense categories."""
    return EXPENSE_TAXONOMY


# ---------------------------------------------------------------------------
# Excel-Export  (must be before /{tx_id} wildcard)
# ---------------------------------------------------------------------------

@router.get("/export/xlsx")
async def export_transactions_xlsx(
    year: Optional[int] = None,
    account_number: Optional[str] = None,
):
    """Exportiert alle Buchungen (optional gefiltert nach Jahr/Konto) als Excel-Datei."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    db = get_db()
    query: dict = {}
    if year:
        from datetime import date
        query["booking_date"] = {"$gte": date(year, 1, 1).isoformat(), "$lte": date(year, 12, 31).isoformat()}
    if account_number:
        query["account_number"] = account_number

    docs = await db.transactions.find(query).sort("booking_date", 1).to_list(None)

    wb = Workbook()
    ws = wb.active
    label = f"Buchungen {year or 'Alle'}"
    ws.title = label[:31]

    PRIMARY = "1A7A5E"
    headers = [
        "Datum", "Konto", "Buchungstext", "Gegenseite", "Verwendungszweck",
        "Betrag", "Währung", "Haupttyp", "Untertyp", "WE", "Fehlbuchung",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=PRIMARY)
    alt_fill    = PatternFill("solid", fgColor="E6F4EF")
    right_align = Alignment(horizontal="right")
    center      = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    for r_idx, doc in enumerate(docs, 2):
        betrag = doc.get("amount")
        row = [
            doc.get("booking_date", ""),
            doc.get("account_number", ""),
            doc.get("booking_text", ""),
            doc.get("counterparty_name", ""),
            doc.get("purpose", ""),
            betrag,
            doc.get("currency", "EUR"),
            doc.get("haupttyp", ""),
            doc.get("untertyp", ""),
            doc.get("owner_unit", ""),
            "Ja" if doc.get("is_fehlbuchung") else "",
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.font = Font(size=9)
            if r_idx % 2 == 0:
                cell.fill = alt_fill
            if col == 6 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = right_align

    col_widths = [12, 10, 20, 28, 40, 12, 8, 18, 22, 8, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Buchungen_{year or 'Alle'}"
    if account_number:
        filename += f"_{account_number}"
    filename += ".xlsx"

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{tx_id}")
async def get_transaction(tx_id: str):
    db = get_db()
    doc = await db.transactions.find_one({"_id": ObjectId(tx_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return _fix_id(doc)


@router.delete("/{tx_id}", status_code=204)
async def delete_transaction(tx_id: str):
    db = get_db()
    result = await db.transactions.delete_one({"_id": ObjectId(tx_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")


# ---------------------------------------------------------------------------
# Kategorie-Zuweisung (manual override)
# ---------------------------------------------------------------------------

class KategorieUpdate(BaseModel):
    haupttyp: Optional[str] = None   # Hauptkategorie  (None = zurücksetzen)
    untertyp: Optional[str] = None   # Unterkategorie  (None = zurücksetzen)


@router.patch("/{tx_id}/kategorie")
async def set_kategorie(tx_id: str, data: KategorieUpdate):
    """Manually assign or clear the Haupttyp/Untertyp for a transaction."""
    if data.haupttyp and data.untertyp:
        if not is_valid_kategorie(data.haupttyp, data.untertyp):
            raise HTTPException(
                status_code=422,
                detail=f"'{data.untertyp}' ist kein gültiger Untertyp von '{data.haupttyp}'",
            )
    db = get_db()
    update = {
        "haupttyp": data.haupttyp,
        "untertyp": data.untertyp,
    }
    result = await db.transactions.update_one(
        {"_id": ObjectId(tx_id)}, {"$set": update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    doc = await db.transactions.find_one({"_id": ObjectId(tx_id)})
    return _fix_id(doc)


# ---------------------------------------------------------------------------
# Retroactive auto-classification of uncategorised ausgabe transactions
# ---------------------------------------------------------------------------

@router.post("/reclassify", status_code=200)
async def reclassify_expenses():
    """
    Run the auto-classifier over ALL ausgabe transactions that currently have
    no category set and update them in-place.
    Returns the number of transactions that were updated.
    """
    db = get_db()
    cursor = db.transactions.find({
        "transaction_type": "ausgabe",
        "haupttyp": None,
        "is_fehlbuchung": {"$ne": True},
    })

    updated = 0
    async for tx in cursor:
        ht, ut = classify_expense(
            tx.get("counterparty_name") or "",
            tx.get("purpose") or "",
        )
        if ht:
            await db.transactions.update_one(
                {"_id": tx["_id"]},
                {"$set": {"haupttyp": ht, "untertyp": ut}},
            )
            updated += 1

    return {"updated": updated}


class FehlbuchungUpdate(BaseModel):
    is_fehlbuchung: bool
    fehlbuchung_note: Optional[str] = None


@router.patch("/{tx_id}/fehlbuchung")
async def set_fehlbuchung(tx_id: str, data: FehlbuchungUpdate):
    """Mark or unmark a transaction as an erroneous booking (Fehlbuchung).
    Fehlbuchungen are excluded from all Abrechnungen and analytics.
    """
    db = get_db()
    update: dict = {"is_fehlbuchung": data.is_fehlbuchung}
    if data.fehlbuchung_note is not None:
        update["fehlbuchung_note"] = data.fehlbuchung_note
    if not data.is_fehlbuchung:
        update["fehlbuchung_note"] = None
    result = await db.transactions.update_one(
        {"_id": ObjectId(tx_id)}, {"$set": update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    doc = await db.transactions.find_one({"_id": ObjectId(tx_id)})
    return _fix_id(doc)
